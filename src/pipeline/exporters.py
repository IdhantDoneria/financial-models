"""Report exporters: PDF, Excel workbook, Google Doc.

All three consume an :class:`~src.pipeline.runner.AnalysisReport` and write a
self-contained file the user can share. PDF and XLSX are pure Python (reportlab
+ openpyxl). Google Docs requires OAuth credentials the user provides; the
function raises a clear :class:`ModelError` if they're missing rather than
silently failing.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from ..base_model import ModelError, ValidationError
from .runner import AnalysisReport

_TITLE_STYLE = "TitleStyle"


# --------------------------------------------------------------------------- #
# PDF
# --------------------------------------------------------------------------- #
def export_pdf(report: AnalysisReport, output_path: str | Path) -> Path:
    """Write ``report`` as a multi-page PDF via reportlab.

    Layout:
        * Cover — company header, mode, timestamp, extracted-data table.
        * Summary — one-row-per-model results table.
        * Assumptions — every key market/model parameter with rationale.
        * Per-model detail — full ``calculate()`` output.

    Args:
        report: The completed analysis to render.
        output_path: Destination ``.pdf`` file.

    Returns:
        The resolved output path.

    Raises:
        ModelError: If reportlab is not installed.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import LETTER
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
    except ImportError as exc:  # pragma: no cover
        raise ModelError("reportlab is required for PDF export: pip install reportlab") from exc

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(output_path), pagesize=LETTER, title="Financial Models Report",
        leftMargin=0.7 * inch, rightMargin=0.7 * inch,
        topMargin=0.7 * inch, bottomMargin=0.7 * inch,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(_TITLE_STYLE, parent=styles["Title"],
                           textColor=colors.HexColor("#1e3a8a"), fontSize=22)
    sub = ParagraphStyle("Sub", parent=styles["Heading2"],
                        textColor=colors.HexColor("#3f6fd6"), fontSize=13)
    body = styles["BodyText"]
    story: list = []

    company = report.company.company_name or "Financial Analysis Report"
    story.append(Paragraph(company, title))
    story.append(Paragraph(
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} · "
        f"Mode: <b>{report.mode.upper()}</b>", body))
    story.append(Spacer(1, 14))

    # Extracted financials.
    story.append(Paragraph("Extracted financials", sub))
    story.append(_kv_table(report.company.to_dict(), colors))
    story.append(Spacer(1, 12))

    # Assumption summary.
    story.append(Paragraph("Assumptions used", sub))
    story.append(_kv_table(report.assumptions.market_context, colors))
    if report.assumptions.rationale:
        story.append(Spacer(1, 6))
        for (model, param), reason in report.assumptions.rationale.items():
            story.append(Paragraph(f"<b>{model} · {param}:</b> {reason}", body))
    story.append(Spacer(1, 12))

    # Results summary.
    story.append(Paragraph("Model results (summary)", sub))
    story.append(_dataframe_table(report.summary_frame(), colors))
    story.append(PageBreak())

    # Per-model detail.
    for name, res in report.results.items():
        story.append(Paragraph(name, sub))
        story.append(_kv_table(res, colors))
        story.append(Spacer(1, 12))
    for name, err in report.errors.items():
        story.append(Paragraph(f"{name} — <font color='red'>FAILED</font>", sub))
        story.append(Paragraph(err, body))

    doc.build(story)
    return output_path


def _kv_table(mapping: dict, colors_mod: Any):
    """Two-column key/value table used across the PDF sections."""
    from reportlab.platypus import Table, TableStyle

    rows = [[str(k), _fmt(v)] for k, v in mapping.items() if v is not None]
    if not rows:
        rows = [["(no data)", ""]]
    table = Table(rows, colWidths=[2.6 * 72, 4.2 * 72])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors_mod.HexColor("#eef2ff")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors_mod.HexColor("#1e3a8a")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors_mod.HexColor("#c7d2fe")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    return table


def _dataframe_table(df, colors_mod: Any):
    """Render a pandas DataFrame as a reportlab table."""
    from reportlab.platypus import Table, TableStyle

    data = [list(df.columns)] + df.astype(str).values.tolist()
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors_mod.HexColor("#1e3a8a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors_mod.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors_mod.HexColor("#94a3b8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors_mod.white, colors_mod.HexColor("#f8fafc")]),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    return table


def _fmt(v: Any) -> str:
    """Human-readable formatting for numbers/lists/objects in report tables."""
    if isinstance(v, float):
        if abs(v) >= 1e6:
            return f"{v:,.0f}"
        return f"{v:.4f}"
    if isinstance(v, list) and v and isinstance(v[0], (int, float)):
        return ", ".join(f"{x:,.2f}" for x in v[:8])
    if isinstance(v, dict):
        return ", ".join(f"{k}={_fmt(val)}" for k, val in list(v.items())[:6])
    return str(v)


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
def export_xlsx(report: AnalysisReport, output_path: str | Path) -> Path:
    """Write ``report`` as an Excel workbook (openpyxl).

    Sheets:
        * ``Summary`` — company header + model results table.
        * ``Extracted`` — every scraped financial figure.
        * ``Assumptions`` — market context + rationale.
        * One sheet per model — full ``calculate()`` output.

    Args:
        report: The completed analysis to render.
        output_path: Destination ``.xlsx`` file.

    Returns:
        The resolved output path.

    Raises:
        ModelError: If openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:  # pragma: no cover
        raise ModelError("openpyxl is required for XLSX export: pip install openpyxl") from exc

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Summary sheet.
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = report.company.company_name or "Financial Analysis Report"
    ws["A1"].font = Font(size=16, bold=True, color="1E3A8A")
    ws["A2"] = f"Generated {datetime.now():%Y-%m-%d %H:%M} — Mode: {report.mode.upper()}"
    ws.append([])
    ws.append(["Model", "Headline result", "Status"])
    for col in range(1, 4):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font, cell.fill = header_font, header_fill
    for _, row in report.summary_frame().iterrows():
        ws.append([row["Model"], row["Headline result"], row["Status"]])
    for col_letter, width in zip("ABC", (32, 24, 16)):
        ws.column_dimensions[col_letter].width = width

    # Extracted financials sheet.
    ws = wb.create_sheet("Extracted")
    _dump_kv(ws, report.company.to_dict(), header_font, header_fill, left)

    # Assumptions.
    ws = wb.create_sheet("Assumptions")
    _dump_kv(ws, report.assumptions.market_context, header_font, header_fill, left)
    if report.assumptions.rationale:
        ws.append([])
        ws.append(["Rationale"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        for (model, param), reason in report.assumptions.rationale.items():
            ws.append([f"{model} — {param}", reason])

    # Per-model sheets.
    for name, res in report.results.items():
        sheet_name = name[:28].replace("/", "-")
        ws = wb.create_sheet(sheet_name)
        _dump_kv(ws, res, header_font, header_fill, left)
    for name, err in report.errors.items():
        sheet_name = f"ERR-{name[:24]}".replace("/", "-")
        ws = wb.create_sheet(sheet_name)
        ws.append(["Error", err])

    wb.save(output_path)
    return output_path


def _dump_kv(ws: Any, mapping: dict, header_font, header_fill, left) -> None:
    """Write a key/value mapping as a two-column sheet with a styled header."""
    ws.append(["Field", "Value"])
    for col in (1, 2):
        cell = ws.cell(row=1, column=col)
        cell.font, cell.fill = header_font, header_fill
    for k, v in mapping.items():
        ws.append([str(k), _fmt(v)])
        ws.cell(row=ws.max_row, column=2).alignment = left
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 60


# --------------------------------------------------------------------------- #
# Google Docs
# --------------------------------------------------------------------------- #
def export_google_doc(
    report: AnalysisReport,
    *,
    credentials_path: str | Path | None = None,
    token_path: str | Path | None = None,
    document_title: str | None = None,
) -> str:
    """Create a Google Doc with the report content and return its URL.

    Requires a Google Cloud OAuth 2.0 desktop-app client (`credentials.json`)
    downloaded from `console.cloud.google.com`. On first run a browser window
    prompts the user to authorise the Docs API; the resulting token is cached
    to ``token_path`` for subsequent runs.

    Args:
        report: The completed analysis to render.
        credentials_path: Path to ``credentials.json``. Defaults to
            ``~/.config/financial-models/credentials.json``.
        token_path: Where to cache the OAuth token. Defaults to
            ``~/.config/financial-models/token.json``.
        document_title: Title for the created doc (default: company name).

    Returns:
        Editable Google Docs URL.

    Raises:
        ModelError: If the Google client libraries or credentials are missing.
        ValidationError: On malformed credential paths.
    """
    try:
        from google.auth.transport.requests import Request
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from googleapiclient.discovery import build
    except ImportError as exc:  # pragma: no cover
        raise ModelError(
            "Google API client libraries missing. Install: "
            "google-api-python-client google-auth google-auth-oauthlib"
        ) from exc

    scopes = ["https://www.googleapis.com/auth/documents"]
    creds_path = Path(credentials_path) if credentials_path else \
        Path.home() / ".config" / "financial-models" / "credentials.json"
    tok_path = Path(token_path) if token_path else \
        Path.home() / ".config" / "financial-models" / "token.json"

    creds = None
    if tok_path.exists():
        creds = Credentials.from_authorized_user_file(str(tok_path), scopes)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not creds_path.exists():
                raise ModelError(
                    f"Google OAuth credentials not found at {creds_path}. See "
                    "docs/google_docs_setup.md for a 3-step setup guide."
                )
            flow = InstalledAppFlow.from_client_secrets_file(str(creds_path), scopes)
            creds = flow.run_local_server(port=0)
        tok_path.parent.mkdir(parents=True, exist_ok=True)
        tok_path.write_text(creds.to_json())

    service = build("docs", "v1", credentials=creds)
    title = document_title or (report.company.company_name or "Financial Analysis Report")
    doc = service.documents().create(body={"title": title}).execute()
    doc_id = doc["documentId"]
    body_text = _render_plain_text(report)
    service.documents().batchUpdate(
        documentId=doc_id,
        body={"requests": [{"insertText": {"location": {"index": 1}, "text": body_text}}]},
    ).execute()
    return f"https://docs.google.com/document/d/{doc_id}/edit"


def _render_plain_text(report: AnalysisReport) -> str:
    """Serialise a report to plain text for the Google Doc body."""
    lines: list[str] = []
    lines.append(report.company.company_name or "Financial Analysis Report")
    lines.append(f"Generated {datetime.now():%Y-%m-%d %H:%M}  ·  Mode: {report.mode.upper()}")
    lines.append("")
    lines.append("EXTRACTED FINANCIALS")
    for k, v in report.company.to_dict().items():
        if v is not None:
            lines.append(f"  {k}: {_fmt(v)}")
    lines.append("")
    lines.append("ASSUMPTIONS USED")
    for k, v in report.assumptions.market_context.items():
        lines.append(f"  {k}: {_fmt(v)}")
    for (model, param), reason in report.assumptions.rationale.items():
        lines.append(f"  · {model} — {param}: {reason}")
    lines.append("")
    lines.append("MODEL RESULTS")
    for name, res in report.results.items():
        lines.append(f"\n== {name} ==")
        for k, v in res.items():
            lines.append(f"  {k}: {_fmt(v)}")
    for name, err in report.errors.items():
        lines.append(f"\n== {name} — FAILED ==\n  {err}")
    return "\n".join(lines) + "\n"
